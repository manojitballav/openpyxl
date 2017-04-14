from openpyxl import Workbook
import openpyxl

wb = Workbook()
ws = wb.create_sheet("Sheet1",0)
wb.save("new.xlsx")

xfile = openpyxl.load_workbook('new.xlsx')
sheet = xfile.get_sheet_by_name('Sheet1')
for row in range(1,5):
	for column in range(1,5):
		ws.cell(row, column , value = 3)
		wb.save('new.xlsx')

print "Rewrite successful"
