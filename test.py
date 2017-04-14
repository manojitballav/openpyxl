from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
import time,openpyxl
from openpyxl import Workbook


print "This is for Redmi 3S prime from Flipkart"
print time.ctime()
driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.maximize_window()

wb = Workbook()
ws = wb.create_sheet("Sheet1",0)
wb.save("new.xlsx")

r = 0

for val in range(1,6):
	val = str(val)
	xfile = openpyxl.load_workbook('new.xlsx')
	sheet = xfile.get_sheet_by_name('Sheet1')
	driver.get('https://www.flipkart.com/redmi-3s-prime-gold-32-gb/product-reviews/itmehbejabfb46rg?page='+val+'&pid=MOBEKWZYSHHJNWGZ')
	if(val == '1'):
		#Click element for the Certified Users
		driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/div/div[2]/div[2]/div[1]/div[5]/div').click()
	#Loop for the elements in the page
	for item in range(1,11):
		item = str(item)
		try:
			#checking the presence of date element
			date = driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/div/div[2]/div[2]/div[2]/div['+item+']/div/div/div[3]/div[1]/div/p[3]')
		except Exception:
			print "Review date elements not populated\t" + str(time.ctime())
			time.sleep(20)
			driver.refresh()
			time.sleep(5)
		try:
			#checking for the "Read more button"
			but = driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/div/div[2]/div[2]/div[2]/div['+item+']/div/div/div[2]/div/span/span/span')
		except Exception:
			print "Some error in finding the read more button\t" + str(time.ctime())
			pass
		try:
			para = driver.find_element_by_xpath('//*[@id="container"]/div/div[2]/div/div/div[2]/div[2]/div[2]/div['+item+']/div/div/div[2]/div/div')
		except Exception:
			print "Review elements not found for: " +val+ " page \t" + str(time.ctime())
			pass
		try:
			but.click()
		except Exception:
			pass
		try:
			if(len(date.text)>1):
				print r
				ws.cell(row = r, column = 0).value = date.text
				ws.cell(row = r, column = 1).value = para.text
				# worksheet.write(row,col,date.text)
				# worksheet.write(row,col+1,para.text)
				r = r + 1
				wb.save('new.xlsx')

			else:
				print "Some problem here at row: " +str(row) + str(time.ctime())
		except Exception:
			pass
	print "Data Scraped from page: "+val
driver.quit()
print "Done" 
print time.ctime()	
