from openpyxl import Workbook
import openpyxl

wb = Workbook()
ws = wb.create_sheet("Sheet1",0)
wb.save("new.xlsx")

xfile = openpyxl.load_workbook('new.xlsx')
sheet = xfile.get_sheet_by_name('Sheet1')
for i in range(1,5):
	for j in range(1,5):
		ws.cell(row = i, column = j, value = j)
		wb.save('new.xlsx')

print "Rewrite successful"
